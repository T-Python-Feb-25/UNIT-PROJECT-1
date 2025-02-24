import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import sqlite3
from colorama import Fore

DB_NAME = "safety_tracker.db"

def export_to_excel():
    """
    Exports risk data from the SQLite database to an Excel file with formatting.

    This function:
    - Retrieves data from the "risks" table in the SQLite database.
    - Saves the data in an Excel file named 'risk_data.xlsx'.
    - Applies formatting to the header and highlights "High" risk levels.

    Returns:
        None

    Example:
        export_to_excel()
        # Generates 'risk_data.xlsx' with formatted risk data.
    """
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM risks")
    data = cursor.fetchall()
    conn.close()

    if data:
        df = pd.DataFrame(data, columns=["ID", "Risk Type", "Level", "Location", "Date"])
        excel_filename = "risk_data.xlsx"
        df.to_excel(excel_filename, index=False, engine="openpyxl")

        wb = load_workbook(excel_filename)
        ws = wb.active

        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=5):
            for cell in col:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
            for cell in row:
                if cell.value == "High":
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                elif cell.value == "Medium":
                    cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                    cell.font = Font(bold=True, color="000000")
                elif cell.value == "Low":
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    cell.font = Font(bold=True, color="000000")

        wb.save(excel_filename)

        print(Fore.GREEN + "✔ Data exported successfully with formatting to 'risk_data.xlsx'!")
    else:
        print(Fore.YELLOW + "⚠ No data available to export.")
